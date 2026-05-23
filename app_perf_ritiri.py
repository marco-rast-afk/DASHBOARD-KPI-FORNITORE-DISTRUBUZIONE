# app.py — Dashboard Performance SDA — versione Streamlit web + Supabase
import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
from datetime import date, timedelta
import io

from core import (
    leggi_file_corrieri,
    aggrega_filiale,
    calcola_tariffa,
    FASCE_DEFAULT,
    ottieni_fasce_filiale,
)

st.set_page_config(
    page_title="Dashboard Performance SDA",
    page_icon="📦",
    layout="wide",
    initial_sidebar_state="expanded",
)

COLORI_FILIALI = ["#3b82f6", "#22c55e", "#a855f7", "#f59e0b", "#14b8a6", "#ef4444", "#ec4899", "#f97316"]

LAYOUT_DARK = dict(
    plot_bgcolor="#181c24", paper_bgcolor="#0f1117", font_color="#f1f5f9",
    legend=dict(bgcolor="#1e2330", bordercolor="#2a3045"),
    margin=dict(l=0, r=0, t=30, b=0),
)

st.markdown("""
<style>
    .kpi-box {
        background: #1e2330;
        border: 1px solid #2a3045;
        border-radius: 10px;
        padding: 16px 18px 12px;
        text-align: center;
    }
    .kpi-val  { font-size: 2rem; font-weight: 700; margin: 0; line-height: 1.1; }
    .kpi-lbl  { font-size: 0.78rem; color: #94a3b8; margin-top: 4px; }
    div[data-testid="stDataFrame"] { font-size: 0.85rem; }
    section[data-testid="stSidebar"] { background: #181c24; }
</style>
""", unsafe_allow_html=True)

def kpi_card(label: str, value: str, color: str = "#3b82f6"):
    st.markdown(
        f'<div class="kpi-box">'
        f'<p class="kpi-val" style="color:{color}">{value}</p>'
        f'<p class="kpi-lbl">{label}</p>'
        f'</div>',
        unsafe_allow_html=True,
    )

def fmt_eur(v: float) -> str:
    return f"€ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

def fmt_n(v: float) -> str:
    return f"{int(v):,}".replace(",", ".")

def colore_filiale(filiali: list, nome: str) -> str:
    try:
        return COLORI_FILIALI[filiali.index(nome) % len(COLORI_FILIALI)]
    except ValueError:
        return "#3b82f6"

# ── SUPABASE ──────────────────────────────────────────────────
from supabase import create_client

@st.cache_resource
def get_supabase():
    url = st.secrets["SUPABASE_URL"]
    key = st.secrets["SUPABASE_KEY"]
    return create_client(url, key)

def get_progetto() -> str:
    """Legge il nome del progetto dai secrets Streamlit."""
    return st.secrets["PROGETTO"]

def carica_da_supabase() -> dict | None:
    """Legge i record del progetto corrente da Supabase."""
    sb       = get_supabase()
    progetto = get_progetto()
    rows     = []
    offset   = 0
    while True:
        chunk = (
            sb.table("performance_corrieri")
            .select("*")
            .eq("progetto", progetto)          # ← filtra per progetto
            .range(offset, offset + 999)
            .execute()
            .data
        )
        rows.extend(chunk)
        if len(chunk) < 1000:
            break
        offset += 1000

    if not rows:
        return None

    # Ricostruisce la struttura: {filiale: {data: {giro: {...}}}}
    dati: dict = {}
    for r in rows:
        fil  = r["filiale"]
        d    = date.fromisoformat(r["data"])
        giro = r["giro"]
        dati.setdefault(fil, {}).setdefault(d, {})[giro] = {
            "lv_af":   r["lv_af"],
            "lv_ok":   r["lv_ok"],
            "lv_rit":  r["lv_rit"],
            "stop_ok": r["stop_ok"],
            "stop_rit": r["stop_rit"],
            "ldv_tot": r["lv_ok"] + r["lv_rit"],
        }
    return dati

def importa_su_supabase(dati_nuovi: dict) -> int:
    """
    Fa upsert dei dati su Supabase (accoda senza duplicare).
    La tabella deve avere UNIQUE (progetto, filiale, data, giro).
    """
    sb       = get_supabase()
    progetto = get_progetto()
    records  = []
    for filiale, giorni in dati_nuovi.items():
        for d, giri in giorni.items():
            for giro, v in giri.items():
                records.append({
                    "progetto": progetto,          # ← aggiunto
                    "filiale":  filiale,
                    "data":     d.isoformat(),
                    "giro":     str(giro),
                    "lv_af":    int(v.get("lv_af",   0)),
                    "lv_ok":    int(v.get("lv_ok",   0)),
                    "lv_rit":   int(v.get("lv_rit",  0)),
                    "stop_ok":  int(v.get("stop_ok", 0)),
                    "stop_rit": int(v.get("stop_rit", 0)),
                })

    # Upsert a blocchi di 500 record
    for i in range(0, len(records), 500):
        sb.table("performance_corrieri").upsert(
            records[i:i + 500],
            on_conflict="progetto,filiale,data,giro"  # ← aggiunto progetto
        ).execute()

    return len(records)

# ══════════════════════════════════════════════════════════════
# RITIRI — Costanti dominio
# ══════════════════════════════════════════════════════════════
_STATI_ESITO = [
    "CLIENTE TRASFERITO", "INDIRIZZO ERRATO / INCOMPLETO",
    "RITIRO GIA' EFFET. IN GIORNATA", "NULLA DA RITIRARE",
    "MERCE NON PRONTA", "CLIENTE ASSENTE", "SPEDIZIONE RITIRATA",
]
_TIPOLOGIE_POSTE = ["OFFERTA UNICA", "OFFERTA UNICA VOLUMINOSO",
                    "POSTE DELIVERY WEB", "POSTE IOINVIO"]
_TIPOLOGIE_SDA   = ["P.F. ASSICURATO ABBONATO", "P.F. VOLUMINOSI",
                    "P.A. VOLUMINOSI", "SDA PORTO ASSEGNATO", "SDA PORTO FRANCO"]
_TIPOLOGIE_FISSI = ["RITIRO FISSO", "SERVIZI A CALENDARIO"]
_TIPOLOGIE_UPS   = ["UPS", "UPS RS"]
_TUTTE_TIP       = _TIPOLOGIE_POSTE + _TIPOLOGIE_SDA + _TIPOLOGIE_FISSI + _TIPOLOGIE_UPS

_COLS_SORGENTE = [
    "Id Ritiro", "Codice Prenotazione", "Canale", "Tipologia",
    "Ragione Sociale", "Telefono", "Indirizzo", "Data",
    "Filiale", "Giro", "Stato Lavorazione", "Stato Ritiro",
    "LV Ritirate", "Note Corriere", "Id Ritiro UPS",
]


def _leggi_ritiri(uploaded_file) -> list[dict]:
    """Legge un file Excel/CSV di ritiri e restituisce lista di dict."""
    import openpyxl, csv as _csv
    nome = uploaded_file.name.lower()
    rows = []

    if nome.endswith(".xls"):
        try:
            import xlrd
            wb = xlrd.open_workbook(file_contents=uploaded_file.read())
            ws = wb.sheets()[0]
            for ri in range(ws.nrows):
                row = []
                for ci in range(ws.ncols):
                    cell = ws.cell(ri, ci)
                    if cell.ctype == 0:
                        row.append(None)
                    elif cell.ctype == 2:
                        v = cell.value
                        row.append(int(v) if v == int(v) else v)
                    else:
                        row.append(str(cell.value).strip())
                rows.append(row)
        except ImportError:
            raise ImportError("Installa xlrd per leggere file .xls")
    elif nome.endswith(".csv"):
        raw = uploaded_file.read()
        for enc in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                text = raw.decode(enc)
                delim = ";" if text.count(";") > text.count(",") else ","
                rows = list(_csv.reader(text.splitlines(), delimiter=delim))
                break
            except Exception:
                continue
    else:
        wb = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
        ws = wb.active
        rows = [list(r) for r in ws.iter_rows(min_row=1, values_only=True)]

    if not rows:
        return []

    hdr_row = next(
        (i for i, r in enumerate(rows)
         if any(str(v or "").strip() == "Id Ritiro" for v in r)), 0)
    headers = [str(h or "").strip() for h in rows[hdr_row]]
    col_idx = {h: i for i, h in enumerate(headers) if h in _COLS_SORGENTE}

    result = []
    for row in rows[hdr_row + 1:]:
        if not any(row):
            continue
        rec = {col: str(row[i] or "").strip()
               for col, i in col_idx.items() if i < len(row)}
        if rec.get("Id Ritiro"):
            result.append(rec)
    return result


def _calcola_ritiri(righe: list[dict]) -> dict:
    """Calcola KPI aggregati dai ritiri — stessa logica del .pyw."""
    from collections import defaultdict
    valide = [r for r in righe
              if r.get("Stato Ritiro", "").strip().upper() != "ANNULLATO"]

    pivot = defaultdict(lambda: defaultdict(int))
    for r in valide:
        tip   = r.get("Tipologia", "").strip()
        stato = r.get("Stato Ritiro", "").strip()
        if stato in _STATI_ESITO:
            pivot[tip][stato] += 1

    def _sezione(tips, col_lav):
        ris = {}
        for tip in tips:
            tot = sum(pivot[tip].values())
            lav = pivot[tip].get(col_lav, 0)
            ris[tip] = {"tot": tot, "lav": lav}
        tot_tot = sum(v["tot"] for v in ris.values())
        tot_lav = sum(v["lav"] for v in ris.values())
        pct     = tot_lav / tot_tot if tot_tot else 0
        return ris, tot_tot, tot_lav, pct

    poste, tot_p, lav_p, pct_p = _sezione(_TIPOLOGIE_POSTE, "SPEDIZIONE RITIRATA")
    sda,   tot_s, rit_s, pct_s = _sezione(_TIPOLOGIE_SDA,   "SPEDIZIONE RITIRATA")
    fissi, tot_f, lav_f, pct_f = _sezione(_TIPOLOGIE_FISSI, "SPEDIZIONE RITIRATA")
    ups,   tot_u, rit_u, pct_u = _sezione(_TIPOLOGIE_UPS,   "SPEDIZIONE RITIRATA")

    n_ldv = sum(int(r.get("LV Ritirate", "0") or 0)
                for r in valide
                if r.get("Stato Ritiro", "").strip() == "SPEDIZIONE RITIRATA")

    return {
        "pivot":    dict(pivot),
        "valide":   valide,
        "totale":   len(righe),
        "n_ldv":    n_ldv,
        "tot_p": tot_p, "lav_p": lav_p, "pct_p": pct_p, "poste": poste,
        "tot_s": tot_s, "rit_s": rit_s, "pct_s": pct_s, "sda":   sda,
        "tot_f": tot_f, "lav_f": lav_f, "pct_f": pct_f, "fissi": fissi,
        "tot_u": tot_u, "rit_u": rit_u, "pct_u": pct_u, "ups":   ups,
    }


def _pubblica_ritiri(righe: list[dict], calc: dict,
                     nome_file: str, filiale: str,
                     data_rif: str) -> int:
    """
    1. DELETE ritiri_dettaglio (project=RITIRI, filiale) — rimuove il giorno precedente
    2. INSERT ritiri_dettaglio  (righe correnti)
    3. INSERT ritiri_storico    (KPI aggregati)
    Ritorna numero righe inserite nel dettaglio.
    """
    sb = get_supabase()

    # 1. Cancella dettaglio precedente
    sb.table("ritiri_dettaglio") \
      .delete() \
      .eq("project", "RITIRI") \
      .eq("filiale", filiale) \
      .execute()

    # 2. Inserisci dettaglio (batch 200)
    rows_det = [{
        "project":              "RITIRI",
        "filiale":              filiale,
        "data_riferimento":     data_rif,
        "id_ritiro":            r.get("Id Ritiro", ""),
        "codice_prenotazione":  r.get("Codice Prenotazione", ""),
        "canale":               r.get("Canale", ""),
        "tipologia":            r.get("Tipologia", ""),
        "ragione_sociale":      r.get("Ragione Sociale", ""),
        "telefono":             r.get("Telefono", ""),
        "indirizzo":            r.get("Indirizzo", ""),
        "data_ritiro":          r.get("Data", ""),
        "filiale_riga":         r.get("Filiale", ""),
        "giro":                 r.get("Giro", ""),
        "stato_lavorazione":    r.get("Stato Lavorazione", ""),
        "stato_ritiro":         r.get("Stato Ritiro", ""),
        "lv_ritirate":          r.get("LV Ritirate", ""),
        "note_corriere":        r.get("Note Corriere", ""),
        "id_ritiro_ups":        r.get("Id Ritiro UPS", ""),
    } for r in righe]

    BATCH = 200
    for i in range(0, len(rows_det), BATCH):
        sb.table("ritiri_dettaglio").insert(rows_det[i:i + BATCH]).execute()

    # 3. Inserisci storico KPI
    c = calc
    n_ass  = sum(c["pivot"].get(t, {}).get("CLIENTE ASSENTE",   0) for t in _TUTTE_TIP)
    n_np   = sum(c["pivot"].get(t, {}).get("MERCE NON PRONTA",  0) for t in _TUTTE_TIP)
    n_nulla= sum(c["pivot"].get(t, {}).get("NULLA DA RITIRARE", 0) for t in _TUTTE_TIP)
    n_ann  = c["totale"] - len(c["valide"])
    sb.table("ritiri_storico").insert({
        "project":          "RITIRI",
        "filiale":          filiale,
        "data_riferimento": data_rif,
        "nome_file":        nome_file,
        "totale":           c["totale"],
        "valide":           len(c["valide"]),
        "ritirati":         c["lav_p"] + c["rit_s"] + c["lav_f"] + c["rit_u"],
        "ldv":              c["n_ldv"],
        "assenti":          n_ass,
        "non_pronti":       n_np,
        "nulla":            n_nulla,
        "annullati":        n_ann,
        "tot_p": c["tot_p"], "lav_p": c["lav_p"], "pct_p": round(c["pct_p"], 6),
        "tot_s": c["tot_s"], "rit_s": c["rit_s"], "pct_s": round(c["pct_s"], 6),
        "tot_f": c["tot_f"], "lav_f": c["lav_f"], "pct_f": round(c["pct_f"], 6),
        "tot_u": c["tot_u"], "rit_u": c["rit_u"], "pct_u": round(c["pct_u"], 6),
        "has_detail": True,
    }).execute()

    return len(rows_det)


@st.cache_data(ttl=120)
def _carica_storico_ritiri(filiale: str) -> list[dict]:
    sb = get_supabase()
    rows = (sb.table("ritiri_storico")
              .select("*")
              .eq("project",  "RITIRI")
              .eq("filiale",  filiale)
              .order("data_riferimento", desc=True)
              .limit(90)
              .execute()
              .data)
    return rows or []


@st.cache_data(ttl=120)
def _carica_dettaglio_ritiri(filiale: str) -> list[dict]:
    sb = get_supabase()
    rows = (sb.table("ritiri_dettaglio")
              .select("*")
              .eq("project", "RITIRI")
              .eq("filiale", filiale)
              .order("id",   desc=False)
              .execute()
              .data)
    return rows or []

# ── SESSION STATE ──────────────────────────────────────────────
if "dati"    not in st.session_state: st.session_state.dati    = None
if "date_da" not in st.session_state: st.session_state.date_da = None
if "date_a"  not in st.session_state: st.session_state.date_a  = None

# ── SIDEBAR ───────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## 📦 Dashboard Performance")
    progetto_attivo = st.secrets.get("PROGETTO", "—")
    st.markdown(f"🗂 **Progetto:** `{progetto_attivo}`")
    st.markdown("---")

    # ── SEZIONE IMPORTA DATI ──
    st.markdown("### 📥 Importa Dati")
    st.caption("Carica un file Excel: i dati vengono accodati a quelli esistenti su Supabase.")

    uploaded = st.file_uploader("Seleziona file Excel", type=["xlsx", "xls"])
    if uploaded:
        if st.button("⬆ Importa su Supabase", use_container_width=True, type="primary"):
            with st.spinner("Lettura e importazione in corso..."):
                try:
                    dati_nuovi = leggi_file_corrieri(uploaded)
                    n_righe = importa_su_supabase(dati_nuovi)
                    st.success(f"✅ {n_righe} record importati!")
                    st.session_state.dati = None  # forza ricaricamento dal DB
                    st.rerun()
                except Exception as e:
                    st.error(f"Errore importazione: {e}")

    st.markdown("---")

    # ── CARICA DATI DA SUPABASE ──
    if st.session_state.dati is None:
        with st.spinner("Caricamento dati da Supabase..."):
            try:
                st.session_state.dati = carica_da_supabase()
                if st.session_state.dati:
                    st.success(f"✅ {len(st.session_state.dati)} filiali caricate")
                else:
                    st.warning("Nessun dato presente. Importa un file Excel.")
            except Exception as e:
                st.error(f"Errore connessione Supabase: {e}")

    if st.button("🔄 Ricarica da Supabase", use_container_width=True):
        st.session_state.dati = None
        st.rerun()

    st.markdown("---")

    # ── FILTRO DATE ──
    if st.session_state.dati:
        tutte_date = sorted({d for fil in st.session_state.dati.values() for d in fil})
        if tutte_date:
            st.markdown("### 📅 Periodo")
            col1, col2 = st.columns(2)
            with col1:
                date_da = st.date_input("Dal", value=tutte_date[0], min_value=tutte_date[0], max_value=tutte_date[-1])
            with col2:
                date_a = st.date_input("Al", value=tutte_date[-1], min_value=tutte_date[0], max_value=tutte_date[-1])
            st.session_state.date_da = date_da
            st.session_state.date_a  = date_a
            c1, c2, c3 = st.columns(3)
            with c1:
                if st.button("Oggi"):
                    st.session_state.date_da = st.session_state.date_a = tutte_date[-1]
                    st.rerun()
            with c2:
                if st.button("7gg"):
                    st.session_state.date_a  = tutte_date[-1]
                    st.session_state.date_da = max(tutte_date[0], tutte_date[-1] - timedelta(days=6))
                    st.rerun()
            with c3:
                if st.button("Tutto"):
                    st.session_state.date_da = tutte_date[0]
                    st.session_state.date_a  = tutte_date[-1]
                    st.rerun()

if not st.session_state.dati:
    st.info("👈 Importa un file Excel dalla barra laterale, oppure attendi il caricamento da Supabase.")
    st.stop()

dati    = st.session_state.dati
filiali = sorted(dati.keys())
date_da = st.session_state.date_da
date_a  = st.session_state.date_a

tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs([
    "📊 Panoramica",
    "🏢 Dettaglio Filiale",
    "📋 Tutti i Giri",
    "📅 Giornaliero",
    "💶 Tariffa",
    "📦 Ritiri",
])

# ══════════════════════════════════════════════════════════════
# TAB 1 — PANORAMICA
# ══════════════════════════════════════════════════════════════
with tab1:
    st.markdown("### Panoramica Multi-Filiale")
    riepilogo = []
    for fil in filiali:
        agg, _, _ = aggrega_filiale(dati[fil], date_da, date_a)
        if agg:
            riepilogo.append({"filiale": fil, **agg})

    if riepilogo:
        df_riep = pd.DataFrame(riepilogo)
        tot_af  = df_riep["tot_lv_af"].sum()
        tot_ok  = df_riep["tot_lv_ok"].sum()
        tot_rit = df_riep["tot_lv_rit"].sum()
        tot_giri_giorni = sum(
            sum(len(giri) for d, giri in dati[f].items()
                if (date_da is None or d >= date_da) and (date_a is None or d <= date_a))
            for f in filiali)
        prod_complessiva_media = (tot_ok + tot_rit) / tot_giri_giorni if tot_giri_giorni > 0 else 0.0

        c1, c2, c3, c4, c5 = st.columns(5)
        with c1: kpi_card("Filiali attive",       str(len(riepilogo)),             "#3b82f6")
        with c2: kpi_card("Tot LV Affidate",       fmt_n(tot_af),                  "#3b82f6")
        with c3: kpi_card("Tot LV Ok",             fmt_n(tot_ok),                  "#22c55e")
        with c4: kpi_card("Tot LV Ritiro",         fmt_n(tot_rit),                 "#a855f7")
        with c5: kpi_card("Prod. Media Corrieri",  f"{prod_complessiva_media:.1f}", "#f59e0b")

        st.markdown("---")

        # GRAFICO 1 — Produttività LDV OK+RIT per filiale
        st.markdown("#### Produttività Media Corrieri (LDV OK+RIT) per Filiale")
        fig_prod = go.Figure()
        for i, row in df_riep.iterrows():
            col = colore_filiale(filiali, row["filiale"])
            fig_prod.add_trace(go.Bar(
                x=[row["filiale"]], y=[round(row["media_prod"], 1)],
                marker_color=col, name=row["filiale"],
                text=[f"{row['media_prod']:.1f}"], textposition="outside",
                showlegend=False,
            ))
        fig_prod.update_layout(**LAYOUT_DARK, height=300, xaxis=dict(gridcolor="#2a3045"), yaxis=dict(gridcolor="#2a3045", title="Media Giornaliera Pezzi per Corriere"))
        st.plotly_chart(fig_prod, use_container_width=True)

        # GRAFICO 2 — LV Ok vs LV Ritiro per filiale
        st.markdown("#### LV Ok vs LV Ritiro per Filiale")
        fig_lv = go.Figure()
        fig_lv.add_trace(go.Bar(
            name="LV Ok",
            x=df_riep["filiale"], y=df_riep["tot_lv_ok"],
            marker_color="#22c55e",
            text=df_riep["tot_lv_ok"].apply(lambda v: fmt_n(v)),
            textposition="outside",
        ))
        fig_lv.add_trace(go.Bar(
            name="LV Ritiro",
            x=df_riep["filiale"], y=df_riep["tot_lv_rit"],
            marker_color="#a855f7",
            text=df_riep["tot_lv_rit"].apply(lambda v: fmt_n(v)),
            textposition="outside",
        ))
        fig_lv.update_layout(**LAYOUT_DARK, barmode="group", height=300, xaxis=dict(gridcolor="#2a3045"), yaxis=dict(gridcolor="#2a3045"))
        st.plotly_chart(fig_lv, use_container_width=True)

        # TABELLA RIEPILOGATIVA
        st.markdown("#### Tabella Riepilogativa Filiali")
        df_tab = df_riep[["filiale", "n_giorni", "tot_lv_af", "tot_lv_ok", "tot_lv_rit", "media_prod"]].copy()
        df_tab.columns = ["Filiale", "Giorni Periodo", "Tot LV AF", "Tot LV Ok", "Tot LV Rit", "Prod. Media Corriere"]
        df_tab["Tot LV AF"]  = df_tab["Tot LV AF"].apply(fmt_n)
        df_tab["Tot LV Ok"]  = df_tab["Tot LV Ok"].apply(fmt_n)
        df_tab["Tot LV Rit"] = df_tab["Tot LV Rit"].apply(fmt_n)
        df_tab["Prod. Media Corriere"] = df_tab["Prod. Media Corriere"].apply(lambda x: f"{x:.1f}")
        st.dataframe(df_tab, use_container_width=True, hide_index=True)

# ══════════════════════════════════════════════════════════════
# TAB 2 — DETTAGLIO FILIALE
# ══════════════════════════════════════════════════════════════
with tab2:
    st.markdown("### Dettaglio Singola Filiale")
    fil_sel = st.selectbox("Seleziona filiale", filiali, key="sel_fil")
    agg, giornate, per_giro = aggrega_filiale(dati[fil_sel], date_da, date_a)

    if agg:
        c1, c2, c3, c4, c5 = st.columns(5)
        with c1: kpi_card("Giorni Attivi",        str(agg["n_giorni"]),        "#94a3b8")
        with c2: kpi_card("Tot LV Affidate",       fmt_n(agg["tot_lv_af"]),    "#3b82f6")
        with c3: kpi_card("Tot LV Ok",             fmt_n(agg["tot_lv_ok"]),    "#22c55e")
        with c4: kpi_card("Tot LV Rit",            fmt_n(agg["tot_lv_rit"]),   "#a855f7")
        with c5: kpi_card("Prod. Media Corrieri",  f"{agg['media_prod']:.1f}", "#f59e0b")

        st.markdown("---")

        # GRAFICO — Andamento giornaliero LV Ok e LV Rit
        st.markdown("#### Andamento Giornaliero LV Ok e LV Ritiro")
        giorni_data = []
        for d in sorted(giornate):
            lv_ok_d  = sum(v["lv_ok"]  for v in giornate[d].values())
            lv_rit_d = sum(v["lv_rit"] for v in giornate[d].values())
            giorni_data.append({"data": d, "lv_ok": lv_ok_d, "lv_rit": lv_rit_d})
        df_giorni = pd.DataFrame(giorni_data)

        if not df_giorni.empty:
            fig_trend = go.Figure()
            fig_trend.add_trace(go.Scatter(
                x=df_giorni["data"], y=df_giorni["lv_ok"],
                name="LV Ok", line=dict(color="#22c55e", width=2),
                fill="tozeroy", fillcolor="rgba(34,197,94,0.1)",
                mode="lines+markers", marker=dict(size=5),
            ))
            fig_trend.add_trace(go.Scatter(
                x=df_giorni["data"], y=df_giorni["lv_rit"],
                name="LV Rit", line=dict(color="#a855f7", width=2),
                mode="lines+markers", marker=dict(size=5),
            ))
            fig_trend.update_layout(**LAYOUT_DARK, height=280, xaxis=dict(gridcolor="#2a3045"), yaxis=dict(gridcolor="#2a3045"))
            st.plotly_chart(fig_trend, use_container_width=True)

        # GRAFICO — Produttività media per giro (barre orizzontali)
        st.markdown("#### Produttività Media per Giro (LDV OK+RIT)")
        if per_giro:
            giri_sorted = sorted(per_giro.items())
            fig_giri = go.Figure(go.Bar(
                y=[f"Giro {g}" for g, _ in giri_sorted],
                x=[round(v["ldv_tot"], 1) for _, v in giri_sorted],
                orientation="h",
                marker_color=[colore_filiale(filiali, fil_sel)] * len(giri_sorted),
                text=[f"{v['ldv_tot']:.1f}" for _, v in giri_sorted],
                textposition="outside",
            ))
            fig_giri.update_layout(**LAYOUT_DARK, height=max(250, len(giri_sorted) * 28), xaxis=dict(gridcolor="#2a3045", title="LDV OK+RIT medi/giorno"), yaxis=dict(gridcolor="#2a3045", autorange="reversed"))
            fig_giri.update_layout(margin=dict(l=0, r=60, t=20, b=0))
            st.plotly_chart(fig_giri, use_container_width=True)

        # TABELLA GIRI
        st.markdown("#### Rendimento dei singoli Giri (Medie Giornaliere del Periodo)")
        if per_giro:
            rows_giro = [{
                "Giro":                              g,
                "Giorni Presenza":                   v["n"],
                "LV AF (Media)":                     f"{v['lv_af']:.1f}",
                "LV Ok (Media)":                     f"{v['lv_ok']:.1f}",
                "LV Rit (Media)":                    f"{v['lv_rit']:.1f}",
                "Stop Ok (Media)":                   f"{v['stop_ok']:.1f}",
                "Stop Rit (Media)":                  f"{v['stop_rit']:.1f}",
                "Prod Media Giorno (LV OK + RIT)":   f"{v['ldv_tot']:.1f}",
            } for g, v in sorted(per_giro.items())]
            st.dataframe(pd.DataFrame(rows_giro), use_container_width=True, hide_index=True)

# ══════════════════════════════════════════════════════════════
# TAB 3 — TUTTI I GIRI
# ══════════════════════════════════════════════════════════════
with tab3:
    st.markdown("### Elenco Completo Multi-Filiale di tutti i Giri")
    righe_tutti = []
    for fil in filiali:
        _, _, pg = aggrega_filiale(dati[fil], date_da, date_a)
        if pg:
            for giro, v in sorted(pg.items()):
                righe_tutti.append({
                    "Filiale":                           fil,
                    "Giro":                              giro,
                    "Giorni":                            v["n"],
                    "LV AF (Media)":                     round(v["lv_af"],   1),
                    "LV Ok (Media)":                     round(v["lv_ok"],   1),
                    "LV Rit (Media)":                    round(v["lv_rit"],  1),
                    "Stop Ok (Media)":                   round(v["stop_ok"], 1),
                    "Stop Rit (Media)":                  round(v["stop_rit"],1),
                    "Prod Media Giorno (LV OK + RIT)":   round(v["ldv_tot"], 1),
                })

    if righe_tutti:
        df_tutti = pd.DataFrame(righe_tutti)
        fil_filter = st.multiselect("Filtra per filiale", filiali, default=filiali, key="filter_tutti")
        df_tutti_f = df_tutti[df_tutti["Filiale"].isin(fil_filter)]

        # GRAFICO — Confronto produttività per giro tra filiali selezionate
        if not df_tutti_f.empty and len(fil_filter) > 0:
            st.markdown("#### Confronto Produttività per Giro tra Filiali")
            fig_conf = go.Figure()
            for fil in fil_filter:
                df_f = df_tutti_f[df_tutti_f["Filiale"] == fil]
                fig_conf.add_trace(go.Bar(
                    name=fil,
                    x=df_f["Giro"].astype(str),
                    y=df_f["Prod Media Giorno (LV OK + RIT)"],
                    marker_color=colore_filiale(filiali, fil),
                ))
            fig_conf.update_layout(**LAYOUT_DARK, barmode="group", height=320, xaxis=dict(gridcolor="#2a3045", title="Giro"), yaxis=dict(gridcolor="#2a3045", title="LDV OK+RIT medi/giorno"))
            st.plotly_chart(fig_conf, use_container_width=True)

        st.dataframe(df_tutti_f, use_container_width=True, hide_index=True)

        # Download CSV
        csv = df_tutti_f.to_csv(index=False, sep=";").encode("utf-8-sig")
        st.download_button("⬇ Scarica CSV", data=csv,
                           file_name="tutti_giri.csv", mime="text/csv")

# ══════════════════════════════════════════════════════════════
# TAB 4 — GIORNALIERO
# ══════════════════════════════════════════════════════════════
with tab4:
    st.markdown("### Dettaglio Giornaliero per Filiale")

    fil_giorno = st.selectbox(
        "Seleziona filiale",
        filiali,
        key="fil_giorno"
    )

    _, giornate_g, _ = aggrega_filiale(
        dati[fil_giorno],
        date_da,
        date_a
    )

    if giornate_g:

        date_sel = st.selectbox(
            "Seleziona data",
            sorted(giornate_g.keys(), reverse=True),
            format_func=lambda d: d.strftime("%d/%m/%Y")
        )

        giri_day = giornate_g[date_sel]

        # KPI giornata
        lv_af_g  = sum(v.get("lv_af",0) for v in giri_day.values())
        lv_ok_g  = sum(v.get("lv_ok",0) for v in giri_day.values())
        lv_rit_g = sum(v.get("lv_rit",0) for v in giri_day.values())
        stop_ok  = sum(v.get("stop_ok",0) for v in giri_day.values())

        tot_ldv    = lv_ok_g + lv_rit_g
        n_giri     = len(giri_day)
        prod_media = tot_ldv / n_giri if n_giri > 0 else 0
        rdc        = (lv_ok_g / lv_af_g * 100) if lv_af_g > 0 else 0

        c1, c2, c3, c4, c5, c6, c7 = st.columns(7)

        with c1:
            kpi_card("LV Affidate",       fmt_n(lv_af_g),          "#3b82f6")
        with c2:
            kpi_card("LV Ok",             fmt_n(lv_ok_g),          "#22c55e")
        with c3:
            kpi_card("LV Ritiro",         fmt_n(lv_rit_g),         "#a855f7")
        with c4:
            kpi_card("Stop Ok",           fmt_n(stop_ok),          "#14b8a6")
        with c5:
            kpi_card("Volume Totale LDV", fmt_n(tot_ldv),          "#94a3b8")
        with c6:
            kpi_card("Prod. Media",       f"{prod_media:.1f}",     "#f59e0b")
        with c7:
            kpi_card("RDC",               f"{rdc:.1f}%",           "#ef4444")

        st.markdown("---")

        righe_giorno = []

        for g, v in sorted(giri_day.items()):
            righe_giorno.append({
                "Giro":                          g,
                "LV AFF":                        int(v.get("lv_af",  0)),
                "LV OK":                         int(v.get("lv_ok",  0)),
                "LV RIT":                        int(v.get("lv_rit", 0)),
                "STOP OK":                       int(v.get("stop_ok",0)),
                "STOP RIT":                      int(v.get("stop_rit",0)),
                "Produttività (LV OK + RIT)":    int(v.get("ldv_tot",0)),
            })

        st.markdown("#### LV Ok e LV Ritiro per Giro")

        fig_day = go.Figure()

        fig_day.add_trace(go.Bar(
            y=[f"Giro {r['Giro']}" for r in righe_giorno],
            x=[r["LV OK"] for r in righe_giorno],
            name="LV Ok",
            orientation="h",
            marker_color="#22c55e",
        ))

        fig_day.add_trace(go.Bar(
            y=[f"Giro {r['Giro']}" for r in righe_giorno],
            x=[r["LV RIT"] for r in righe_giorno],
            name="LV Rit",
            orientation="h",
            marker_color="#a855f7",
        ))

        fig_day.update_layout(
            **LAYOUT_DARK,
            barmode="stack",
            height=max(250, len(righe_giorno) * 30),
            yaxis=dict(gridcolor="#2a3045", autorange="reversed"),
            xaxis=dict(gridcolor="#2a3045")
        )

        st.plotly_chart(fig_day, use_container_width=True)

        st.dataframe(
            pd.DataFrame(righe_giorno),
            use_container_width=True,
            hide_index=True
        )

    else:
        st.warning("Nessun dato disponibile nel periodo selezionato.")

# ══════════════════════════════════════════════════════════════
# TAB 5 — TARIFFA
# ══════════════════════════════════════════════════════════════
with tab5:
    st.markdown("### Calcolo Fatturato a Scaglioni Progressivi")
    fil_tar = st.selectbox("Seleziona filiale per tariffazione", filiali, key="fil_tar")
    _, giornate_t, _ = aggrega_filiale(dati[fil_tar], date_da, date_a)

    if giornate_t:
        fasce_attive = ottieni_fasce_filiale(fil_tar)

        with st.expander(f"Ispeziona scaglioni attivi per la filiale: {fil_tar}", expanded=False):
            for idx, f in enumerate(fasce_attive):
                st.write(f"Scaglione {idx+1} ➔ Da: {f['da']} a {f['a']} LDV | Tariffa: € {f['prezzo']:.3f}")

        righe_tar, tot_v, tot_f, med_f = calcola_tariffa(giornate_t, fasce_attive)

        c1, c2, c3, c4 = st.columns(4)
        with c1: kpi_card("Giorni",                   str(len(righe_tar)),  "#94a3b8")
        with c2: kpi_card("Volume Totale (LV OK+RIT)", fmt_n(tot_v),        "#3b82f6")
        with c3: kpi_card("Fatturato Totale Stimato",  fmt_eur(tot_f),      "#22c55e")
        with c4: kpi_card("Fatturato Medio / Giorno",  fmt_eur(med_f),      "#a855f7")

        st.markdown("---")

        # GRAFICO — Fatturato giornaliero
        st.markdown("#### Fatturato Giornaliero")
        df_tar = pd.DataFrame(righe_tar)
        fig_tar = go.Figure(go.Bar(
            x=df_tar["data"],
            y=df_tar["fatturato"],
            marker_color="#22c55e",
            text=[fmt_eur(v) for v in df_tar["fatturato"]],
            textposition="outside",
        ))
        fig_tar.update_layout(**LAYOUT_DARK, height=300, xaxis=dict(gridcolor="#2a3045"), yaxis=dict(gridcolor="#2a3045", tickprefix="€ "))
        fig_tar.update_layout(margin=dict(l=0, r=0, t=30, b=0))
        st.plotly_chart(fig_tar, use_container_width=True)

        # GRAFICO — Volume LDV giornaliero
        st.markdown("#### Volume LDV (OK+RIT) Giornaliero")
        fig_vol = go.Figure(go.Bar(
            x=df_tar["data"],
            y=df_tar["volume"],
            marker_color="#3b82f6",
            text=[fmt_n(v) for v in df_tar["volume"]],
            textposition="outside",
        ))
        fig_vol.update_layout(**LAYOUT_DARK, height=260, xaxis=dict(gridcolor="#2a3045"), yaxis=dict(gridcolor="#2a3045", title="LDV"))
        fig_vol.update_layout(margin=dict(l=0, r=0, t=20, b=0))
        st.plotly_chart(fig_vol, use_container_width=True)

        st.markdown("#### Dettaglio giornaliero scaglioni")
        st.dataframe(df_tar, use_container_width=True, hide_index=True)
    else:
        st.warning("Nessun dato disponibile nel periodo selezionato.")

# ══════════════════════════════════════════════════════════════
# TAB 6 — RITIRI
# ══════════════════════════════════════════════════════════════
with tab6:
    st.markdown("### 📦 Calcolo & Storico Ritiri")

    # ── Filiale ritiri dai secrets ─────────────────────────────
    filiale_ritiri = st.secrets.get("FILIALE_RITIRI", "")
    if not filiale_ritiri:
        st.warning(
            "⚠️ Aggiungi `FILIALE_RITIRI` nei secrets Streamlit "
            "(es. `FILIALE_RITIRI = \"AP\"`) per abilitare questa sezione."
        )
        st.stop()

    col_sx, col_dx = st.columns([1, 2])

    # ─────────────────────────────────────────────────────────
    # COLONNA SX — Carica file + pubblica
    # ─────────────────────────────────────────────────────────
    with col_sx:
        st.markdown("#### 📥 Carica File Ritiri")
        up_ritiri = st.file_uploader(
            "File giornaliero (xlsx / xls / csv)",
            type=["xlsx", "xls", "csv"],
            key="up_ritiri",
        )

        if up_ritiri:
            data_rif_input = st.date_input(
                "Data di riferimento",
                value=date.today(),
                key="data_rif_ritiri",
            )

            if st.button("⬆️ Calcola & Pubblica su Dashboard",
                         type="primary", use_container_width=True):
                with st.spinner("Lettura, calcolo e pubblicazione in corso..."):
                    try:
                        righe_r = _leggi_ritiri(up_ritiri)
                        if not righe_r:
                            st.error("Nessun dato trovato nel file.")
                        else:
                            calc_r = _calcola_ritiri(righe_r)
                            n_pub  = _pubblica_ritiri(
                                righe_r, calc_r,
                                up_ritiri.name,
                                filiale_ritiri,
                                data_rif_input.isoformat(),
                            )
                            _carica_storico_ritiri.clear()
                            _carica_dettaglio_ritiri.clear()
                            st.success(
                                f"✅ Pubblicati {n_pub} righe dettaglio "
                                f"e KPI per il {data_rif_input.strftime('%d/%m/%Y')}"
                            )
                            st.session_state["ritiri_calc"]  = calc_r
                            st.session_state["ritiri_righe"] = righe_r
                            st.rerun()
                    except Exception as ex:
                        st.error(f"Errore: {ex}")

        # ── KPI calcolati in sessione (dopo upload) ────────────
        if "ritiri_calc" in st.session_state:
            c = st.session_state["ritiri_calc"]
            n_ritirati = c["lav_p"] + c["rit_s"] + c["lav_f"] + c["rit_u"]
            n_ann      = c["totale"] - len(c["valide"])
            n_ass      = sum(c["pivot"].get(t, {}).get("CLIENTE ASSENTE",  0) for t in _TUTTE_TIP)
            n_np       = sum(c["pivot"].get(t, {}).get("MERCE NON PRONTA", 0) for t in _TUTTE_TIP)

            st.markdown("---")
            st.markdown("##### KPI file corrente")
            m1, m2 = st.columns(2)
            with m1: kpi_card("Totale Ritiri",    str(c["totale"]),   "#3b82f6")
            with m2: kpi_card("Ritirati",          str(n_ritirati),   "#22c55e")
            m3, m4 = st.columns(2)
            with m3: kpi_card("LDV Ritirate",     str(c["n_ldv"]),    "#22c55e")
            with m4: kpi_card("Annullati",         str(n_ann),         "#ef4444")
            m5, m6 = st.columns(2)
            with m5: kpi_card("Assenti",           str(n_ass),         "#f59e0b")
            with m6: kpi_card("Merce Non Pronta",  str(n_np),          "#f59e0b")

            st.markdown("##### % per Categoria")
            p1, p2, p3, p4 = st.columns(4)
            with p1: kpi_card("Poste", f"{c['pct_p']:.1%}", "#a855f7")
            with p2: kpi_card("SDA",   f"{c['pct_s']:.1%}", "#3b82f6")
            with p3: kpi_card("Fissi", f"{c['pct_f']:.1%}", "#22c55e")
            with p4: kpi_card("UPS",   f"{c['pct_u']:.1%}", "#f59e0b")

    # ─────────────────────────────────────────────────────────
    # COLONNA DX — Storico / Dettaglio / Pivot
    # ─────────────────────────────────────────────────────────
    with col_dx:
        rtab1, rtab2, rtab3 = st.tabs([
            "📅 Storico KPI",
            "📋 Dettaglio Ultimo Giorno",
            "📊 Pivot Tipologie",
        ])

        # ── Storico KPI ───────────────────────────────────────
        with rtab1:
            if st.button("🔄 Aggiorna Storico", key="btn_refresh_sto"):
                _carica_storico_ritiri.clear()

            storico_rows = _carica_storico_ritiri(filiale_ritiri)

            if not storico_rows:
                st.info("Nessun dato nello storico. Pubblica il primo file per iniziare.")
            else:
                df_sto = pd.DataFrame(storico_rows)

                st.markdown("#### Trend % Ritiro per Categoria")
                df_chart = df_sto.sort_values("data_riferimento")
                fig_trend_r = go.Figure()
                for col_pct, nome, colore in [
                    ("pct_p", "Poste", "#a855f7"),
                    ("pct_s", "SDA",   "#3b82f6"),
                    ("pct_f", "Fissi", "#22c55e"),
                    ("pct_u", "UPS",   "#f59e0b"),
                ]:
                    if col_pct in df_chart.columns:
                        fig_trend_r.add_trace(go.Scatter(
                            x=df_chart["data_riferimento"],
                            y=(df_chart[col_pct].astype(float) * 100).round(1),
                            name=nome,
                            mode="lines+markers",
                            line=dict(color=colore, width=2),
                            marker=dict(size=5),
                        ))
                fig_trend_r.update_layout(
                    **LAYOUT_DARK, height=260,
                    xaxis=dict(gridcolor="#2a3045"),
                    yaxis=dict(gridcolor="#2a3045", ticksuffix="%", range=[0, 105]),
                )
                st.plotly_chart(fig_trend_r, use_container_width=True)

                st.markdown("#### Storico KPI")
                cols_show = ["data_riferimento", "nome_file", "totale", "valide",
                             "ritirati", "ldv", "assenti", "non_pronti",
                             "pct_p", "pct_s", "pct_f", "pct_u"]
                cols_show = [c for c in cols_show if c in df_sto.columns]
                df_sto_tab = df_sto[cols_show].copy()
                for pc in ["pct_p", "pct_s", "pct_f", "pct_u"]:
                    if pc in df_sto_tab.columns:
                        df_sto_tab[pc] = df_sto_tab[pc].apply(
                            lambda v: f"{float(v):.1%}" if v is not None else "—")
                rename_map = {
                    "data_riferimento": "Data", "nome_file": "File",
                    "totale": "Totali",         "valide": "Valide",
                    "ritirati": "Ritirati",     "ldv": "LDV",
                    "assenti": "Assenti",        "non_pronti": "Non Pronti",
                    "pct_p": "% Poste",          "pct_s": "% SDA",
                    "pct_f": "% Fissi",          "pct_u": "% UPS",
                }
                df_sto_tab.rename(columns=rename_map, inplace=True)
                st.dataframe(df_sto_tab, use_container_width=True, hide_index=True)

                csv_sto = df_sto_tab.to_csv(index=False, sep=";").encode("utf-8-sig")
                st.download_button("⬇ Scarica CSV Storico", data=csv_sto,
                                   file_name="storico_ritiri.csv", mime="text/csv")

        # ── Dettaglio Ultimo Giorno ────────────────────────────
        with rtab2:
            st.caption(
                "Righe raw dell'**ultimo file pubblicato**. "
                "Al caricamento successivo vengono automaticamente sostituite."
            )
            if st.button("🔄 Aggiorna Dettaglio", key="btn_refresh_det"):
                _carica_dettaglio_ritiri.clear()

            det_rows = _carica_dettaglio_ritiri(filiale_ritiri)

            if not det_rows:
                st.info("Nessun dettaglio. Pubblica un file per visualizzarlo.")
            else:
                data_det = det_rows[0].get("data_riferimento", "")[:10]
                st.markdown(f"**Riferimento:** `{data_det}` — **{len(det_rows)} righe**")

                fc1, fc2, fc3 = st.columns(3)
                with fc1:
                    stati_u = sorted({r.get("stato_ritiro","") for r in det_rows if r.get("stato_ritiro")})
                    filtro_stato = st.selectbox("Stato Ritiro", ["Tutti"] + stati_u, key="filt_stato_det")
                with fc2:
                    tipi_u = sorted({r.get("tipologia","") for r in det_rows if r.get("tipologia")})
                    filtro_tipo = st.selectbox("Tipologia", ["Tutte"] + tipi_u, key="filt_tipo_det")
                with fc3:
                    giri_u = sorted({r.get("giro","") for r in det_rows if r.get("giro")},
                                    key=lambda x: int(x) if x.isdigit() else 999)
                    filtro_giro = st.selectbox("Giro", ["Tutti"] + giri_u, key="filt_giro_det")

                det_filtrate = [
                    r for r in det_rows
                    if (filtro_stato == "Tutti" or r.get("stato_ritiro","") == filtro_stato)
                    and (filtro_tipo  == "Tutte" or r.get("tipologia","")   == filtro_tipo)
                    and (filtro_giro  == "Tutti" or r.get("giro","")        == filtro_giro)
                ]
                st.caption(f"{len(det_filtrate)} / {len(det_rows)} righe")

                cols_det = ["tipologia", "ragione_sociale", "stato_ritiro",
                            "giro", "data_ritiro", "lv_ritirate", "note_corriere"]
                df_det = pd.DataFrame([
                    {c: r.get(c,"") for c in cols_det} for r in det_filtrate
                ])
                if not df_det.empty:
                    df_det.rename(columns={
                        "tipologia": "Tipologia", "ragione_sociale": "Ragione Sociale",
                        "stato_ritiro": "Stato Ritiro", "giro": "Giro",
                        "data_ritiro": "Data", "lv_ritirate": "LV",
                        "note_corriere": "Note",
                    }, inplace=True)
                    st.dataframe(df_det, use_container_width=True, hide_index=True)
                    csv_det = df_det.to_csv(index=False, sep=";").encode("utf-8-sig")
                    st.download_button(
                        "⬇ Scarica CSV Dettaglio", data=csv_det,
                        file_name=f"dettaglio_ritiri_{data_det}.csv",
                        mime="text/csv")

        # ── Pivot Tipologie (da sessione) ──────────────────────
        with rtab3:
            if "ritiri_calc" not in st.session_state:
                st.info("Carica un file a sinistra per vedere il pivot.")
            else:
                c = st.session_state["ritiri_calc"]
                pivot = c["pivot"]

                righe_piv = []
                for tip in _TUTTE_TIP:
                    row = {"Tipologia": tip}
                    tot = 0
                    for stato in _STATI_ESITO:
                        v = pivot.get(tip, {}).get(stato, 0)
                        row[stato] = int(v) if v else None
                        tot += v
                    row["TOTALE"] = int(tot) if tot else None
                    if tot:
                        righe_piv.append(row)

                if righe_piv:
                    row_tot = {"Tipologia": "TOTALE COMPLESSIVO"}
                    for stato in _STATI_ESITO:
                        v = sum(pivot.get(t, {}).get(stato, 0) for t in _TUTTE_TIP)
                        row_tot[stato] = int(v) if v else None
                    row_tot["TOTALE"] = sum(
                        v for v in row_tot.values() if isinstance(v, int)) or None
                    righe_piv.append(row_tot)

                    df_piv = pd.DataFrame(righe_piv).fillna("")
                    st.dataframe(df_piv, use_container_width=True, hide_index=True)

                st.markdown("#### % Ritiro per Categoria")
                cat_data = [
                    ("Poste", c["pct_p"], "#a855f7"),
                    ("SDA",   c["pct_s"], "#3b82f6"),
                    ("Fissi", c["pct_f"], "#22c55e"),
                    ("UPS",   c["pct_u"], "#f59e0b"),
                ]
                fig_cat = go.Figure([go.Bar(
                    x=[d[0] for d in cat_data],
                    y=[round(d[1] * 100, 1) for d in cat_data],
                    marker_color=[d[2] for d in cat_data],
                    text=[f"{d[1]:.1%}" for d in cat_data],
                    textposition="outside",
                )])
                fig_cat.update_layout(
                    **LAYOUT_DARK, height=280,
                    yaxis=dict(gridcolor="#2a3045", ticksuffix="%", range=[0, 115]),
                    xaxis=dict(gridcolor="#2a3045"),
                )
                st.plotly_chart(fig_cat, use_container_width=True)
